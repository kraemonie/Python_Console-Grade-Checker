import sys
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

students = {}          # {student_name: student_name}
subjects = {}          # {subject_name: subject_name}
grades = {}            # {student_name: {subject_name: grade}}

def export_to_excel():
    from openpyxl.styles import Font, Alignment, PatternFill

    filename = "grade_report.xlsx"

    df_students = pd.DataFrame({"Student Name": list(students.keys())})
    df_subjects = pd.DataFrame({"Subject Name": list(subjects.keys())})

    all_grade_rows = []
    for student, subdict in grades.items():
        for subject, grade in subdict.items():
            try:
                g = float(grade)
            except:
                g = None
            status = "PASSED" if (g is not None and g >= 75) else "FAILED"
            
            all_grade_rows.append({
                "Student Name": student,
                "Subject": subject,
                "Grade": grade,
                "Status": status
            })
    df_all_grades = pd.DataFrame(all_grade_rows)

    with pd.ExcelWriter(filename, engine="openpyxl", mode="w") as writer:
        
        df_students.to_excel(writer, sheet_name="Students", index=False)
        df_subjects.to_excel(writer, sheet_name="Subjects", index=False)
        df_all_grades.to_excel(writer, sheet_name="All Grades", index=False)

        for subject_name in subjects:
            subject_rows = []
            
            for student_name in students:
                if student_name in grades and subject_name in grades[student_name]:
                    grade_val = grades[student_name][subject_name]
                    
                    try:
                        g = float(grade_val)
                    except:
                        g = None
                    status = "PASSED" if (g is not None and g >= 75) else "FAILED"

                    subject_rows.append({
                        "Student Name": student_name,
                        "Subject": subject_name,
                        "Grade": grade_val,
                        "Status": status
                    })

            df_sub = pd.DataFrame(subject_rows)
            
            if df_sub.empty:
                df_sub = pd.DataFrame(columns=["Student Name", "Subject", "Grade", "Status"])
                
            df_sub.to_excel(writer, sheet_name=subject_name, index=False)


    wb = load_workbook(filename)
    
    
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center")
    green_fill = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid") 
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")   
    white_font = Font(color="FFFFFF") 

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        
        for cell in ws[1]:
            cell.font = header_font
            cell.alignment = center_align

        
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = center_align

        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = length + 4
 
        if sheet_name not in ["Students", "Subjects"]:
            
            
            status_col_idx = None
            for cell in ws[1]:
                if cell.value == "Status":
                    status_col_idx = cell.column - 1 
                    break
            
            if status_col_idx is not None:
                for row in ws.iter_rows(min_row=2):
                    status_cell = row[status_col_idx]
                    
                    if status_cell.value == "PASSED":
                        status_cell.fill = green_fill
                        status_cell.font = white_font 
                    elif status_cell.value == "FAILED":
                        status_cell.fill = red_fill
                        status_cell.font = white_font

    wb.save(filename)
    print(f"\nExcel file saved as: {filename}")

def pause():
    input("\nPress Enter to continue...")


# STUDENTS
def add_student():
    name = input("Enter student name: ").strip()
    if name in students:
        print("\nStudent already exists!")
        return
    students[name] = name
    print(f"\nStudent added successfully: {name}")


def edit_student():
    name = input("Enter student name to edit: ").strip()
    if name not in students:
        print("\nStudent not found!")
        return

    new_name = input("Enter new name: ").strip()

    if name in grades:
        grades[new_name] = grades.pop(name)

    students.pop(name)
    students[new_name] = new_name

    print("\nStudent updated!")


def delete_student():
    name = input("Enter student name to delete: ").strip()
    if name not in students:
        print("\nStudent not found!")
        return

    students.pop(name)
    grades.pop(name, None)

    print("\nStudent deleted!")


def view_students():
    if not students:
        print("\nNo students available.")
        return

    print("\n--- STUDENTS LIST ---")
    for name in students:
        print(name)


# SUBJECTS
def add_subject():
    subject = input("Enter subject name: ").strip()
    if subject in subjects:
        print("\nSubject already exists!")
        return

    subjects[subject] = subject
    print("\nSubject added!")


def edit_subject():
    subject = input("Enter subject name to edit: ").strip()
    if subject not in subjects:
        print("\nSubject not found!")
        return

    new_subject = input("Enter new subject name: ").strip()

    for s in grades:
        if subject in grades[s]:
            grades[s][new_subject] = grades[s].pop(subject)

    subjects.pop(subject)
    subjects[new_subject] = new_subject

    print("\nSubject updated!")


def delete_subject():
    subject = input("Enter subject name to delete: ").strip()
    if subject not in subjects:
        print("\nSubject not found!")
        return

    subjects.pop(subject)

    for s in grades:
        grades[s].pop(subject, None)

    print("\nSubject deleted!")


def view_subjects():
    if not subjects:
        print("No subjects available.")
        return

    print("\n--- SUBJECTS LIST ---")
    for name in subjects:
        print(name)


# GRADES
def pick_student():
    if not students:
        print("\nNo students available!")
        return None
    view_students()
    return input("Enter student name: ").strip()


def pick_subject():
    if not subjects:
        print("\nNo subjects available!")
        return None
    view_subjects()
    return input("Enter subject name: ").strip()


def add_grade():
    student = pick_student()
    if not student or student not in students:
        print("\nInvalid student!")
        return

    if not subjects:
        print("\nNo subjects available! Please add subjects first.")
        return

    print(f"\n--- Entering Grades for: {student} ---")
    print("(Press Enter without typing to skip a subject)\n")

    for subject in subjects:
        current_grade = grades.get(student, {}).get(subject, "N/A")

        new_grade = input(f"   Enter grade for '{subject}' (current: {current_grade}): ").strip()

        if new_grade:
            grades.setdefault(student, {})[subject] = new_grade
        else:
            print(f"   -> Skipped {subject}")

    print("\nDone entering grades for this student.")


def edit_grade():
    student = pick_student()
    if not student or student not in grades:
        print("\nNo grades for this student.")
        return

    subject = pick_subject()
    if not subject or subject not in grades[student]:
        print("\nNo grade for this subject.")
        return

    grade = input("Enter new grade: ").strip()
    grades[student][subject] = grade
    print("\nGrade updated!")


def delete_grade():
    student = pick_student()
    if not student or student not in grades:
        print("\nNo grades for this student.")
        return

    subject = pick_subject()
    if not subject or subject not in grades[student]:
        print("\nGrade not found.")
        return

    grades[student].pop(subject)
    print("\nGrade deleted!")


def view_all_grades():
    if not students:
        print("\nNo students available.")
        return

    print("\n===== ALL STUDENTS' GRADES =====")

    for sName, studentName in students.items():
        print(f"\n--- {studentName} ---")

        if sName not in grades or not grades[sName]:
            print("  No grades available.")
            continue

        for sub, g in grades[sName].items():
            print(f"  {sub}: {g}")


# MENUS
def menu():
    while True:
        print("\n===== GRADE CHECKER SYSTEM =====")
        print("1. Manage Students")
        print("2. Manage Subjects")
        print("3. Manage Grades")
        print("4. Generate Report")
        print("5. Exit")

        choice = input("Select an option: ").strip()

        if choice == "1":
            student_menu()
        elif choice == "2":
            subject_menu()
        elif choice == "3":
            grade_menu()
        elif choice == "4":
            report_menu()
        elif choice == "5":
            print("\nExiting system...")
            sys.exit()
        else:
            print("Invalid choice!")


def student_menu():
    while True:
        print("\n--- MANAGE STUDENTS ---")
        print("1. Add")
        print("2. Edit")
        print("3. Delete")
        print("4. View")
        print("0. Back")

        choice = input("Select: ").strip()

        if choice == "1":
            add_student()
        elif choice == "2":
            edit_student()
        elif choice == "3":
            delete_student()
        elif choice == "4":
            view_students()
        elif choice == "0":
            break
        else:
            print("\nInvalid Choice!")
        pause()


def subject_menu():
    while True:
        print("\n--- MANAGE SUBJECTS ---")
        print("1. Add")
        print("2. Edit")
        print("3. Delete")
        print("4. View")
        print("0. Back")

        choice = input("Select: ").strip()

        if choice == "1":
            add_subject()
        elif choice == "2":
            edit_subject()
        elif choice == "3":
            delete_subject()
        elif choice == "4":
            view_subjects()
        elif choice == "0":
            break
        else:
            print("\nInvalid Choice!")
        pause()


def grade_menu():
    while True:
        print("\n--- MANAGE GRADES ---")
        print("1. Input Grades")
        print("2. View Grades")
        print("0. Back")

        choice = input("Select: ").strip()

        if choice == "1":
            input_grades_menu()
        elif choice == "2":
            view_all_grades()
            pause()
        elif choice == "0":
            break
        else:
            print("\nInvalid Choice!")
            pause()


def report_menu():
    while True:
        print("\n--- GENERATE REPORT ---")
        print("1. Generate Students")
        print("2. Generate Subjects")
        print("3. Generate Grades")
        print("4. Export to Excel")
        print("0. Back")

        choice = input("Select: ").strip()

        if choice == "1":
            view_students()
            pause()
        elif choice == "2":
            view_subjects()
            pause()
        elif choice == "3":
            view_all_grades()
            pause()
        elif choice == "4":
            export_to_excel()
            pause()
        elif choice == "0":
            break
        else:
            print("\nInvalid Choice!")
            pause()


def input_grades_menu():
    while True:
        print("\n--- INPUT GRADES ---")
        print("1. Add (pick subject)")
        print("2. Edit (pick subject)")
        print("3. Delete (pick subject)")
        print("0. Back")

        choice = input("Select: ").strip()

        if choice == "1":
            add_grade()
        elif choice == "2":
            edit_grade()
        elif choice == "3":
            delete_grade()
        elif choice == "0":
            break
        else:
            print("\nInvalid Choice!")

        pause()
        
if __name__ == "__main__":
    menu()
